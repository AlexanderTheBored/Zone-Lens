import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Non-Functional Requirements"

# ── Colour palette (matches the FR document) ──────────────────────────────────
DARK_GREEN   = "2E4A3E"   # header background
CREAM        = "F5F0E8"   # page background / odd-row fill
WHITE        = "FFFFFF"
RUST         = "C0392B"   # accent (matches the rule line in the FR doc)
LINK_GREEN   = "3A7D5A"   # teal-green used for some FR feature names
TEXT_DARK    = "1C1C1C"
TEXT_LIGHT   = "F5F0E8"

# ── NFR data ──────────────────────────────────────────────────────────────────
nfr_data = [
    # ID, Category, Requirement, Metric / Acceptance Criterion, Priority
    ("NFR-01", "Performance",
     "Map Initial Load Time",
     "Interactive map shall render within 3 seconds on a 10 Mbps connection on first load.",
     "High"),
    ("NFR-02", "Performance",
     "Flood-Data Rendering",
     "Flood-risk layer (60 MB GeoJSON) shall finish painting within 5 seconds after toggle on a modern desktop browser.",
     "High"),
    ("NFR-03", "Performance",
     "Chatbot Response Latency",
     "BelaBot shall return a response within 4 seconds for 95 % of queries under normal network conditions.",
     "High"),
    ("NFR-04", "Performance",
     "Map Pan / Zoom Smoothness",
     "Pan and zoom interactions shall maintain ≥ 30 fps on mid-range hardware (e.g., 2019 laptop, integrated GPU).",
     "Medium"),
    ("NFR-05", "Usability",
     "Learnability",
     "A first-time user with no GIS background shall be able to query a zone and flood risk within 2 minutes without external help.",
     "High"),
    ("NFR-06", "Usability",
     "Chatbot Conversation Flow",
     "BelaBot responses shall be written in plain language (≤ Grade 8 reading level) and avoid raw GIS jargon without explanation.",
     "Medium"),
    ("NFR-07", "Usability",
     "Mobile Responsiveness",
     "All UI controls (layer toggles, search, legend, chatbot) shall be fully usable on screens ≥ 360 px wide without horizontal scrolling.",
     "High"),
    ("NFR-08", "Usability",
     "Legend Clarity",
     "Each zone type and flood-risk level shall have a distinct, labelled colour in the legend; no two entries may share the same hue within the same layer group.",
     "Medium"),
    ("NFR-09", "Accessibility",
     "Colour-Blind Support",
     "Zone and flood-risk colour schemes shall pass WCAG 2.1 AA contrast requirements and be distinguishable under Deuteranopia simulation.",
     "Medium"),
    ("NFR-10", "Accessibility",
     "Keyboard Navigation",
     "All interactive controls (search, layer toggles, chatbot input) shall be reachable and operable via keyboard alone.",
     "Medium"),
    ("NFR-11", "Accessibility",
     "Screen-Reader Compatibility",
     "Map layer status changes and chatbot messages shall be announced via ARIA live regions so screen readers can convey updates.",
     "Low"),
    ("NFR-12", "Reliability",
     "Uptime / Availability",
     "The static site shall target 99.5 % monthly uptime; Cloudflare Worker (BelaBot proxy) shall target 99.0 % monthly uptime.",
     "High"),
    ("NFR-13", "Reliability",
     "Graceful Degradation",
     "If the BelaBot API is unavailable, the map shall remain fully functional and display a user-facing error message rather than a blank screen.",
     "High"),
    ("NFR-14", "Reliability",
     "Data Accuracy",
     "Zoning polygons shall faithfully reflect the Mandaue City CLUP 2019–2029 (Aug 2024 revision); discrepancies shall be flagged and corrected within 30 days of discovery.",
     "High"),
    ("NFR-15", "Security",
     "API Key Exposure",
     "Mapbox and BelaBot API keys shall be scoped to authorised domains and rotated at least every 12 months; keys shall never be committed to public repositories.",
     "High"),
    ("NFR-16", "Security",
     "Input Sanitisation",
     "All user-supplied text (search field, chatbot input) shall be sanitised before processing to prevent XSS and prompt-injection attacks.",
     "High"),
    ("NFR-17", "Security",
     "Data Integrity (Read-Only)",
     "No mechanism shall allow end users to modify or delete zoning or flood-risk data; all writes are restricted to authorised maintainers.",
     "High"),
    ("NFR-18", "Security",
     "HTTPS Enforcement",
     "The application shall only be served over HTTPS; any HTTP request shall be redirected to HTTPS automatically.",
     "High"),
    ("NFR-19", "Compatibility",
     "Browser Support",
     "The application shall be fully functional on the two latest stable versions of Chrome, Firefox, Edge, and Safari (desktop and mobile).",
     "High"),
    ("NFR-20", "Compatibility",
     "Mapbox GL JS Version",
     "The app shall remain compatible with Mapbox GL JS v2.x; any upgrade to v3.x shall be tested against all existing layers and controls before deployment.",
     "Medium"),
    ("NFR-21", "Scalability",
     "Concurrent Users",
     "The static-site architecture shall support at least 500 simultaneous users without degraded map or search performance.",
     "Medium"),
    ("NFR-22", "Scalability",
     "Dataset Growth",
     "The data layer architecture shall support addition of new barangay polygons or updated flood surveys without requiring a full page redesign.",
     "Medium"),
    ("NFR-23", "Maintainability",
     "Code Modularity",
     "GeoJSON data files, map-layer definitions, and UI logic shall be separated into distinct modules/files so each can be updated independently.",
     "Medium"),
    ("NFR-24", "Maintainability",
     "Data Update Cycle",
     "Zoning and flood-risk datasets shall be reviewed and updated at least once per year, or within 60 days of an official amendment by Mandaue City.",
     "High"),
    ("NFR-25", "Maintainability",
     "Documentation",
     "A developer README shall document the deployment process, dataset update procedure, and API key rotation steps.",
     "Low"),
    ("NFR-26", "Legal & Compliance",
     "Disclaimer Visibility",
     "A disclaimer stating the data is for planning reference only (not official certification) shall be visible on every session without requiring user action.",
     "High"),
    ("NFR-27", "Legal & Compliance",
     "Data Attribution",
     "All data sources (CLUP, MCDRRMO, Project NOAH) shall be credited in the application's About/Credits section.",
     "Medium"),
    ("NFR-28", "Privacy",
     "No Personal Data Collection",
     "The application shall not collect, store, or transmit personally identifiable information; chatbot queries shall be anonymised before forwarding to the AI provider.",
     "High"),
]

# ── Sheet styling helpers ─────────────────────────────────────────────────────
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def header_fill():
    return PatternFill("solid", fgColor=DARK_GREEN)

def odd_fill():
    return PatternFill("solid", fgColor="F2EDE3")   # warm cream

def even_fill():
    return PatternFill("solid", fgColor=WHITE)

def priority_fill(p):
    colours = {"High": "F9DADA", "Medium": "FFF3CC", "Low": "D6EAD6"}
    return PatternFill("solid", fgColor=colours.get(p, WHITE))

# ── Title row ─────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 14
ws.row_dimensions[2].height = 38

title_cell = ws["A2"]
ws.merge_cells("A2:E2")
title_cell.value = "Non-Functional Requirements"
title_cell.font = Font(name="Calibri", bold=True, size=20, color=DARK_GREEN)
title_cell.alignment = Alignment(horizontal="left", vertical="center")

# Rust rule line (row 3)
ws.row_dimensions[3].height = 4
for col in range(1, 6):
    c = ws.cell(row=3, column=col)
    c.fill = PatternFill("solid", fgColor=RUST)

# ── Column headers (row 4) ────────────────────────────────────────────────────
ws.row_dimensions[4].height = 28
headers = ["ID", "Category", "Requirement", "Description / Acceptance Criterion", "Priority"]
col_widths = [10, 20, 28, 70, 12]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    c = ws.cell(row=4, column=col_idx, value=h)
    c.fill = header_fill()
    c.font = Font(name="Calibri", bold=True, size=11, color=TEXT_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# ── Data rows ─────────────────────────────────────────────────────────────────
ROW_START = 5
for row_offset, (nfr_id, category, req, desc, priority) in enumerate(nfr_data):
    row = ROW_START + row_offset
    ws.row_dimensions[row].height = 42
    row_fill = odd_fill() if row_offset % 2 == 0 else even_fill()

    values = [nfr_id, category, req, desc, priority]
    for col_idx, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.border = border
        c.alignment = Alignment(
            horizontal="center" if col_idx in (1, 2, 5) else "left",
            vertical="center",
            wrap_text=True,
        )
        if col_idx == 5:
            c.fill = priority_fill(priority)
            c.font = Font(
                name="Calibri", size=10, bold=True,
                color=("B22222" if priority == "High"
                       else "7D6000" if priority == "Medium"
                       else "1E5C1E"),
            )
        elif col_idx == 1:
            c.fill = row_fill
            c.font = Font(name="Calibri", size=10, bold=True, color=TEXT_DARK)
        elif col_idx == 3:
            c.fill = row_fill
            c.font = Font(name="Calibri", size=10, color=LINK_GREEN, bold=False)
        else:
            c.fill = row_fill
            c.font = Font(name="Calibri", size=10, color=TEXT_DARK)

# ── Freeze panes below header ─────────────────────────────────────────────────
ws.freeze_panes = "A5"

# ── Sheet tab colour ──────────────────────────────────────────────────────────
ws.sheet_properties.tabColor = DARK_GREEN

# ── Save ──────────────────────────────────────────────────────────────────────
output_path = "/home/user/Zone-Lens/ZoneLens_NFR.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
